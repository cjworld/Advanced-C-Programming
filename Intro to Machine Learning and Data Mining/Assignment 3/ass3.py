import os
import struct
import numpy as np
from numpy import linalg as LA
from array import array
import matplotlib.pyplot as plt


def getSheetNames(excelfile):
    from pandas import ExcelFile
    return (ExcelFile(excelfile)).sheet_names


def writeExcelData(x, excelfile, sheetname, startrow, startcol, columns):
    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook
    df = DataFrame(x)
    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheetname, columns=columns, startrow=startrow-1, startcol=startcol-1, header=False, index=False)
    writer.save()
    writer.close()


def load_mnist(dataset, selected_digits, path=os.path.dirname(os.path.realpath(__file__))):

    # Check training/testing specification. Must be "training" (default) or "testing"
    if dataset == "training":
        filename_digits = os.path.join(path, 'train-images-idx3-ubyte')
        filename_labels = os.path.join(path, 'train-labels-idx1-ubyte')
    elif dataset == "testing":
        filename_digits = os.path.join(path, 't10k-images-idx3-ubyte')
        filename_labels = os.path.join(path, 't10k-labels-idx1-ubyte')
    else:
        raise ValueError("dataset must be 'testing' or 'training'")

    # Import digits data
    digits_file_object = open(filename_digits, 'rb')
    magic_nr, size, rows, cols = struct.unpack(">IIII", digits_file_object.read(16))
    digits_data = array("B", digits_file_object.read())
    digits_file_object.close()

    # Import label data
    labels_file_object = open(filename_labels, 'rb')
    magic_nr, size = struct.unpack(">II", labels_file_object.read(8))
    labels_data = array("B", labels_file_object.read())
    labels_file_object.close()

    # Find indices of selected digits
    indices = [k for k in range(size) if labels_data[k] in selected_digits]
    n = len(indices)

    # Create empty arrays for vector X and T
    X = np.zeros((n, rows * cols), dtype=np.uint8)
    T = np.zeros((n, 1), dtype=np.uint8)

    # Fill vector X from digits_data
    # Fill vector T from labels_data
    for i in range(n):
        X[i] = digits_data[indices[i] * rows * cols:(indices[i] + 1) * rows * cols]
        T[i] = labels_data[indices[i]]

    return X, T


def vector_to_img(*args, **kwargs):
    print kwargs
    # plt.imshow(v.reshape(28, 28), interpolation='None', cmap='gray')
    # plt.axis('off')
    # if show:
    #    plt.show()
    width = kwargs["width"] if kwargs.has_key("width") else 28
    size = 1
    n = len(args)
    fig = plt.figure()
    for i, arg in enumerate(args):
        plt.subplot(1,n,i+1)
        plt.imshow(arg.reshape(width, width), interpolation="None", cmap="gray")
        plt.axis("off")
    fig.tight_layout(pad=0)
    fig.set_size_inches(w=n*size,h=size)
    plt.show()


def PCA(X, U, V):
    # Get recentered matrix Z
    Z = X - U

    # Get principal component matrix
    P = np.dot(Z, V.T)
    # print P
    # scatter_plot(P, T)

    return P


def XZCVPR(X, T, selected_digits):
    # vector_to_img(X[4], X[-1])

    # Get mean matrix U and matrix Z
    U = np.mean(X, axis=0)
    # print U
    # print type(U)
    # print U.shape

    # Get recentered matrix Z
    Z = X - U
    # print Z
    # print np.mean(Z, axis=0)

    # Get covariance matrix C
    C = np.cov(Z, rowvar=False)
    # print C
    # vector_to_img(C, width=784)

    L, V = LA.eigh(C)
    # print L
    # print V

    # row = V[-1, :]
    # col = V[:, -1]
    # print np.dot(C, row) - (L[-1] * row)
    # print np.dot(C, col) - (L[-1] * col)

    L = np.flipud(L)
    V = np.flipud(V.T)
    # row = V[0, :]
    # print np.dot(C, row) - (L[0] * row)
    # print np.allclose(np.dot(C, row), L[0] * row)
    # vector_to_img(V[0:1, :], width=28)
    # vector_to_img(V[1:2, :], width=28)

    # Get principal component matrix
    P = np.dot(Z, V.T)
    # print P
    scatter_plot(P, T, selected_digits=selected_digits)

    # Get recovered matrix
    R = np.dot(P, V)
    # print R-Z

    # Get recovered data matrix
    Xrec = R+U
    # print Xrec-X

    # reduced X
    Xrec = (np.dot(P[:, 0:2], V[0:2, :])) + U
    # print Xrec
    # vector_to_img(X[0], Xrec[0])

    return U, V, P


def scatter_plot(P, T, selected_digits=[5, 6]):
    # For best effect, points should not be drawn in sequence but in random order
    np.random.seed(0)
    randomorder = np.random.permutation(np.arange(len(T)))
    randomorder = np.arange(len(T))

    # Set colors
    cols = np.zeros((len(T), 4))  # Initialize matrix to hold colors
    for i in range(len(T)):
        if T[i] == selected_digits[0]:
            cols[i] = [1, 0, 0, 0.25]  # Negative points are red (with opacity 0.25)
        else:
            cols[i] = [0, 1, 0, 0.25]  # Positive points are green (with opacity 0.25)

    # Draw scatter plot
    fig = plt.figure()
    ax = fig.add_subplot(111, facecolor='black')
    ax.scatter(P[randomorder, 1], P[randomorder, 0], s=5, linewidths=0, facecolors=cols[randomorder, :], marker="o");
    ax.set_aspect('equal')

    plt.gca().invert_yaxis()
    plt.show()


def build_2d_histograms(Prec, T, B, selected_digits):
    histograms = [np.zeros((B, B), dtype=np.int), np.zeros((B, B), dtype=np.int32)]

    mins = np.amin(Prec, axis=0)
    # print "mins: ", mins

    maxs = np.amax(Prec, axis=0)
    # print "maxs: ", maxs

    for i in range(len(Prec)):
        r = int(np.round(np.divide((B - 1) * (Prec[i, 0] - mins[0]), maxs[0] - mins[0], dtype=np.float64)))
        c = int(np.round(np.divide((B - 1) * (Prec[i, 1] - mins[1]), maxs[1] - mins[1], dtype=np.float64)))
        if T[i] == selected_digits[0]:
            histograms[0][r, c] += 1
        else:
            histograms[1][r, c] += 1
    # print histograms
    vector_to_img(histograms[0], width=B)
    vector_to_img(histograms[1], width=B)

    return histograms, maxs, mins


def get_ccps_by_2d_histograms(B, histograms, maxs, mins, n, x):
    r = int(np.round(np.divide((B - 1) * (x[0] - mins[0]), maxs[0] - mins[0], dtype=np.float64)))
    c = int(np.round(np.divide((B - 1) * (x[1] - mins[1]), maxs[1] - mins[1], dtype=np.float64)))

    ccp = [0., 0.]
    ccp[0] = np.divide(histograms[0][r, c], n[0], dtype=np.float64)
    ccp[1] = np.divide(histograms[1][r, c], n[1], dtype=np.float64)

    return ccp


def build_2d_beysian(Prec, T, selected_digits):
    mus = [None, None]
    mus[0] = np.mean(Prec[T[:, 0] == selected_digits[0]], axis=0)
    mus[1] = np.mean(Prec[T[:, 0] == selected_digits[1]], axis=0)

    Sigmas = [None, None]
    Sigmas[0] = np.cov(Prec[T[:, 0] == selected_digits[0]], rowvar=False)
    Sigmas[1] = np.cov(Prec[T[:, 0] == selected_digits[1]], rowvar=False)

    return mus, Sigmas


def get_ccp_by_2d_beysian(x, mu, Sigma):
    # print mu
    d = np.alen(mu)
    # print d
    dfact1 = (2 * np.pi) ** d
    dfact2 = np.linalg.det(Sigma)
    fact = np.divide(1., np.sqrt(dfact1 * dfact2))
    xc = x - mu
    Invsigma = np.linalg.inv(Sigma)
    power = -0.5 * np.dot(np.dot(xc, Invsigma), xc.T)
    return fact * np.exp(power)


def get_ccps_by_2d_beysian(x, mus, Sigmas):
    return [get_ccp_by_2d_beysian(x, mus[0], Sigmas[0]), get_ccp_by_2d_beysian(x, mus[1], Sigmas[1])]


def posterior_probability(ccps, n):
    total = sum(n)
    # print n[0], n[1], total

    eachps = [0.] * len(n)
    # print eachps
    for i in range(len(n)):
        eachps[i] = ccps[i] * np.divide(n[i], total, dtype=np.float64)
    # print eachps

    ttlp = sum(eachps)
    # print ttlp

    pps = [0.] * len(n)
    # print pps
    for i in range(len(n)):
        if ttlp == 0.:
            pps[i] = 0.
        else:
            pps[i] = np.divide(eachps[i], ttlp, dtype=np.float64)
    # print pps

    return pps


def main():
    selected_digits = [6, 9]

    X, T = load_mnist("training", selected_digits)
    X = np.array(X, np.float64)
    T = np.array(T, np.float64)
    # print type(X)
    # print type(T)
    # print X.shape
    # print T.shape

    # Xn = np.array([[72., 101., 94.], [50., 96., 70.], [14., 79., 10.], [8., 70., 1.]], np.float64)
    U, V, P = XZCVPR(X, T, selected_digits)

    n = [0, 0]
    n[0] = len(T[T == selected_digits[0]])
    n[1] = len(T[T == selected_digits[1]])
    # B = int(np.ceil(np.log2(n[0] + n[1]))) + 1
    B = 32
    # print B

    Prec = P[:, 0:2]
    # print Prec[:, 0]
    # print Prec[:, 1]

    histograms, maxs, mins = build_2d_histograms(Prec, T, B, selected_digits)

    mus, Sigmas = build_2d_beysian(Prec, T, selected_digits)

    Xt, Tt = load_mnist("testing", selected_digits)
    Xt = np.array(Xt, np.float64)
    Tt = np.array(Tt, np.float64)
    # print Tt

    # Get principal component matrix of Xt
    Pt = PCA(Xt, U, V)
    # print Pt

    # Get redecued principal component matrix of Pt
    Ptrec = Pt[:, 0:2]
    # print Ptrec[0]
    # print Ptrec[1]

    Tth = np.zeros((len(Tt), 2))
    Ttb = np.zeros((len(Tt), 2))
    for i in range(len(Tt)):
        ccps_by_histograms = get_ccps_by_2d_histograms(B, histograms, maxs, mins, n, Ptrec[i])
        # print ccps_by_histograms
        pps_by_histograms = posterior_probability(ccps_by_histograms, n)
        # print pps_by_histograms
        if pps_by_histograms[0] == pps_by_histograms[1]:
            Tth[i] = (-1., 0.)
        else:
            if pps_by_histograms[0] > pps_by_histograms[1]:
                Tth[i] = (str(selected_digits[0]), pps_by_histograms[0])
            else:
                Tth[i] = (str(selected_digits[1]), pps_by_histograms[1])

        ccps_by_beysian = get_ccps_by_2d_beysian(Ptrec[i], mus, Sigmas)
        # print ccps_by_beysian
        pps_by_beysian = posterior_probability(ccps_by_beysian, n)
        # print pps_by_beysian
        if pps_by_beysian[0] == pps_by_beysian[1]:
            Ttb[i] = ("Undecidable", 0.)
        else:
            if pps_by_beysian[0] > pps_by_beysian[1]:
                Ttb[i] = (str(selected_digits[0]), pps_by_beysian[0])
            else:
                Ttb[i] = (str(selected_digits[1]), pps_by_beysian[1])
    # print Tth
    # print Ttb
    # print len(Tt[Tt == Tth])
    # print len(Tt[Tt == Ttb])

    idxp = 2
    idxn = 1
    # vector_to_img(Xt[idxp], Xt[idxn])

    xp = Xt[idxp]
    xn = Xt[idxn]
    # print xp
    # print xn

    zp = xp - U
    zn = xn - U
    # print zp
    # print zn

    Pp = np.dot(zp, V.T)
    Pn = np.dot(zn, V.T)
    # print Pp
    # print Pn

    Precp = Pp[0:2]
    Precn = Pn[0:2]
    # print Precp
    # print Precn

    Rp = np.dot(Precp, V[0:2, :])
    Rn = np.dot(Precn, V[0:2, :])
    # print Rp
    # print Rn

    Xrecp = Rp + U
    Xrecn = Rn + U
    # print Xrecp
    # print Xrecn

    Thp = Tth[idxp]
    Tbp = Ttb[idxp]
    # print Thp
    # print Tbp

    Thn = Tth[idxn]
    Tbn = Ttb[idxn]
    # print Thn
    # print Tbn

    ch = 0
    cb = 0
    for i in range(len(Tt)):
        # print Tth[i][0], Tt[i][0], Tth[i][0] == Tt[i][0]
        if Tth[i][0] == Tt[i][0]:
            ch += 1
        # print Ttb[i][0], Tt[i][0], Ttb[i][0] == Tt[i][0]
        if Ttb[i][0] == Tt[i][0]:
            cb += 1
    # print len(Tt), ch, cb

    accuracyh = np.divide(ch, len(Tt), dtype=np.float64)
    accuracyb = np.divide(cb, len(Tt), dtype=np.float64)
    # print accuracyh
    # print accuracyb
    
    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook

    excelfile = "/Users/ChiYuChen/Intro to Machine Learning and Data Mining/Assignment 3/Assignment_3_Submission_Template_Copy.xlsx"
    sheetname = "Results"

    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # print [U]
    df = DataFrame([U])
    df.to_excel(writer, sheet_name=sheetname, startrow=1, startcol=1, header=False, index=False)

    # print [V[0, :]]
    df = DataFrame([V[0, :]])
    df.to_excel(writer, sheet_name=sheetname, startrow=2, startcol=1, header=False, index=False)

    # print [V[1, :]]
    df = DataFrame([V[1, :]])
    df.to_excel(writer, sheet_name=sheetname, startrow=3, startcol=1, header=False, index=False)

    # print [n[0]]
    df = DataFrame([n[0]])
    df.to_excel(writer, sheet_name=sheetname, startrow=5, startcol=1, header=False, index=False)

    # print [n[1]]
    df = DataFrame([n[1]])
    df.to_excel(writer, sheet_name=sheetname, startrow=6, startcol=1, header=False, index=False)

    # print [mus[0]]
    df = DataFrame([mus[0]])
    df.to_excel(writer, sheet_name=sheetname, startrow=8, startcol=1, header=False, index=False)

    # print [mus[1]]
    df = DataFrame([mus[1]])
    df.to_excel(writer, sheet_name=sheetname, startrow=9, startcol=1, header=False, index=False)

    # print Sigmas[0]
    df = DataFrame(Sigmas[0])
    df.to_excel(writer, sheet_name=sheetname, startrow=11, startcol=1, header=False, index=False)

    # print Sigmas[1]
    df = DataFrame(Sigmas[1])
    df.to_excel(writer, sheet_name=sheetname, startrow=13, startcol=1, header=False, index=False)

    # print maxs[0], mins[0]
    df = DataFrame([[maxs[0], mins[0]]])
    df.to_excel(writer, sheet_name=sheetname, startrow=16, startcol=1, header=False, index=False)

    # print maxs[1], mins[1]
    df = DataFrame([[maxs[1], mins[1]]])
    df.to_excel(writer, sheet_name=sheetname, startrow=17, startcol=1, header=False, index=False)

    # print histograms[0]
    # print histograms[0][0, :]
    df = DataFrame(histograms[0])
    df.to_excel(writer, sheet_name=sheetname, startrow=19, startcol=1, header=False, index=False)

    # print histograms[1]
    # print histograms[1][0, :]
    df = DataFrame(histograms[1])
    df.to_excel(writer, sheet_name=sheetname, startrow=52, startcol=1, header=False, index=False)

    # print xp
    df = DataFrame([xp])
    df.to_excel(writer, sheet_name=sheetname, startrow=87, startcol=1, header=False, index=False)
    # print zp
    df = DataFrame([zp])
    df.to_excel(writer, sheet_name=sheetname, startrow=88, startcol=1, header=False, index=False)
    # print Precp
    df = DataFrame([Precp])
    df.to_excel(writer, sheet_name=sheetname, startrow=89, startcol=1, header=False, index=False)
    # print Rp
    df = DataFrame([Rp])
    df.to_excel(writer, sheet_name=sheetname, startrow=90, startcol=1, header=False, index=False)
    # print Xrecp
    df = DataFrame([Xrecp])
    df.to_excel(writer, sheet_name=sheetname, startrow=91, startcol=1, header=False, index=False)

    # print xn
    df = DataFrame([xn])
    df.to_excel(writer, sheet_name=sheetname, startrow=93, startcol=1, header=False, index=False)
    # print zn
    df = DataFrame([zn])
    df.to_excel(writer, sheet_name=sheetname, startrow=94, startcol=1, header=False, index=False)
    # print Precn
    df = DataFrame([Precn])
    df.to_excel(writer, sheet_name=sheetname, startrow=95, startcol=1, header=False, index=False)
    # print Rn
    df = DataFrame([Rn])
    df.to_excel(writer, sheet_name=sheetname, startrow=96, startcol=1, header=False, index=False)
    # print Xrecn
    df = DataFrame([Xrecn])
    df.to_excel(writer, sheet_name=sheetname, startrow=97, startcol=1, header=False, index=False)

    # print selected_digits[0]
    df = DataFrame([selected_digits[0]])
    df.to_excel(writer, sheet_name=sheetname, startrow=101, startcol=1, header=False, index=False)
    # print Thp[0], Thp[1]
    df = DataFrame([[Thp[0], Thp[1]]])
    df.to_excel(writer, sheet_name=sheetname, startrow=102, startcol=1, header=False, index=False)
    # print Tbp[0], Tbp[1]
    df = DataFrame([[Tbp[0], Tbp[1]]])
    df.to_excel(writer, sheet_name=sheetname, startrow=103, startcol=1, header=False, index=False)

    # print selected_digits[1]
    df = DataFrame([selected_digits[1]])
    df.to_excel(writer, sheet_name=sheetname, startrow=105, startcol=1, header=False, index=False)
    # print Thn[0], Thn[1]
    df = DataFrame([[Thn[0], Thn[1]]])
    df.to_excel(writer, sheet_name=sheetname, startrow=106, startcol=1, header=False, index=False)
    # print Tbn[0], Tbn[1]
    df = DataFrame([[Tbn[0], Tbn[1]]])
    df.to_excel(writer, sheet_name=sheetname, startrow=107, startcol=1, header=False, index=False)

    # print accuracyh
    df = DataFrame([[accuracyh]])
    df.to_excel(writer, sheet_name=sheetname, startrow=110, startcol=1, header=False, index=False)
    # print accuracyb
    df = DataFrame([[accuracyb]])
    df.to_excel(writer, sheet_name=sheetname, startrow=111, startcol=1, header=False, index=False)

    writer.save()
    writer.close()


if __name__ == "__main__":
    main()
