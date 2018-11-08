import os
import struct
import numpy as np
from numpy import linalg as LA
from array import array
import matplotlib.pyplot as plt
from scipy.stats import multivariate_normal as mvn


def readExcelSheet1(excelfile):
    from pandas import read_excel
    return (read_excel(excelfile)).values


def readExcelRange(excelfile, sheetname="Sheet1", startrow=1, endrow=1, startcol=1, endcol=1):
    from pandas import read_excel
    values = (read_excel(excelfile, sheetname, header=None)).values
    return values[startrow-1:endrow, startcol-1:endcol]


def readExcel(excelfile, **args):
    if args:
        data = readExcelRange(excelfile, **args)
    else:
        data = readExcelSheet1(excelfile)
    if data.shape == (1, 1):
        return data[0, 0]
    elif (data.shape)[0] == 1:
        return data[0]
    else:
        return data


def getSheetNames(excelfile):
    from pandas import ExcelFile
    return (ExcelFile(excelfile)).sheet_names


def writeExcelData(x, excelfile, sheetname, startrow, startcol, columns):
    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook
    df = DataFrame(x)
    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheetname, columns=columns, startrow=startrow-1, startcol=startcol-1, header=False, index=False)
    writer.save()
    writer.close()

def em_gmm_orig(xs, scales, mus, sigmas, tolerance=0.000001, iterations=10000):
    n, d = xs.shape
    c = len(scales)
    err = 10000.
    rd = 0

    while err > tolerance and rd < iterations:
        print 'Iteration: %d' % rd

        # E-step
        ps = np.zeros((n, c))
        for k in range(c):
            for i in range(n):
                ps[i, k] = scales[k] * mvn.pdf(xs[i], mus[k], sigmas[k])
        # print ps
        for i in range(n):
            # print ps[i, :]
            sumup = np.sum(ps[i, :])
            # print sumup
            for k in range(c):
                ps[i, k] = np.divide(ps[i, k], sumup, dtype=np.float64)
            # print ps[i, :]
        # print ps

        # M-step
        scales = np.zeros(c)
        for k in range(c):
            scales[k] = np.sum(ps[:, k])
            # print scales[k]
        # print scales
        scales /= n
        # print scales

        new_mus = np.zeros((c, d))
        for k in range(c):
            for i in range(n):
                new_mus[k] += ps[i, k] * xs[i]
            new_mus[k] = np.divide(new_mus[k], scales[k] * n, dtype=np.float64)
            # print new_mus[k]
        # print new_mus

        mus_diff = new_mus - mus
        # print new_mus
        # print mus
        # print mus_diff
        # print np.abs(mus_diff)
        # print np.amax(np.abs(mus_diff))
        err = np.amax(np.abs(mus_diff))

        mus = new_mus

        new_sigmas = np.zeros((c, d, d))
        for k in range(c):
            for i in range(n):
                xxs = np.reshape(xs[i], (1, d))
                # print xxs
                z = xxs - mus[k]
                # print z
                # print ps[i, k]
                # print np.dot(z.T, z)
                # print ps[i, k] * np.dot(z.T, z)
                new_sigmas[k] += ps[i, k] * np.dot(z.T, z)
                # print new_sigmas[k]
            new_sigmas[k] = np.divide(new_sigmas[k], scales[k] * n, dtype=np.float64)
        # print new_sigmas

        sigmas = new_sigmas

        rd += 1

    return ps, scales, mus, sigmas


def main():
    excelfile = "/Users/ChiYuChen/Intro to Machine Learning and Data Mining/Assignment 5/Assignment_5_Data_and_Template_Original.xlsx"

    sheetname = "Data"
    xs = np.array(readExcel(excelfile, sheetname=sheetname, startrow=2, endrow=951, endcol=2), dtype=np.float64)
    # print data
    # print data[0]
    # print data[-1]

    dims = xs.shape
    n = dims[0]
    d = dims[1]
    # print d
    c = 3

    # initial random guesses for parameters
    np.random.seed(0)
    scales = np.array([np.divide(1, c, dtype=np.float64)] * c)
    # print scales
    maxs = np.amax(xs, axis=0)
    # print maxs
    mins = np.amin(xs, axis=0)
    # print mins
    mus = np.zeros((c, d))
    for i in range(c):
        for j in range(d):
            mus[i, j] = (maxs[j] - mins[j]) * np.random.random() + mins[j]
            # print j, maxs[j], mins[j], mus[i, j]
    # print mus
    vs = np.var(xs, axis=0)
    # print vars
    sigmas = np.zeros((c, d, d))
    for i in range(c):
        for j in range(d):
            sigmas[i, j, j] = vs[j]
    # print sigmas

    ps, scales, mus, sigmas = em_gmm_orig(xs, scales, mus, sigmas, iterations=10000)

    print scales
    amounts = np.round(scales * n)
    print amounts

    print mus
    idx_male = np.argmax(mus[:, 0])
    idx_children = np.argmin(mus[:, 0])
    idx_female = 3 - idx_male - idx_children
    print idx_male
    print idx_female
    print idx_children
    labelnames = ["F", "F", "F"]
    labelnames[idx_male] = "M"
    labelnames[idx_female] = "F"
    labelnames[idx_children] = "C"
    print labelnames

    print ps
    label_indices = np.argmax(ps, axis=1)
    print label_indices
    labels = np.array([labelnames[label_idx] for label_idx in label_indices])
    print labels
    confs = np.array([ps[idx, label_idx] for idx, label_idx in enumerate(label_indices)])
    print confs
    totals = np.array([amounts[idx_male], amounts[idx_female], amounts[idx_children]])
    print totals

    # In the below plots the white dots represent the observed heights.

    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook

    excelfile = "/Users/ChiYuChen/Intro to Machine Learning and Data Mining/Assignment 5/Assignment_5_Data_and_Template.xlsx"
    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    sheetname = "Results"

    # print labels
    df = DataFrame(labels)
    df.to_excel(writer, sheet_name=sheetname, startrow=1, startcol=0, header=False, index=False)

    # print confs
    df = DataFrame(confs)
    df.to_excel(writer, sheet_name=sheetname, startrow=1, startcol=1, header=False, index=False)
    
    # print totals
    df = DataFrame(totals)
    df.to_excel(writer, sheet_name=sheetname, startrow=1, startcol=5, header=False, index=False)

    writer.save()
    writer.close()


if __name__ == "__main__":
    main()
