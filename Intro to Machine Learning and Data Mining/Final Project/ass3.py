import os
import struct
import numpy as np
from numpy import linalg as LA
from array import array
import matplotlib.pyplot as plt
from ass4 import readExcel



def getSheetNames(excelfile):
    from pandas import ExcelFile
    return (ExcelFile(excelfile)).sheet_names


def writeExcelData(x, excelfile, sheetname, startrow, startcol, columns):
    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook
    df = DataFrame(x)
    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheetname, columns=columns, startrow=startrow-1, startcol=startcol-1, header=False, index=False)
    writer.save()
    writer.close()


def load_mnist(dataset, selected_digits, path=os.path.dirname(os.path.realpath(__file__))):

    # Check training/testing specification. Must be "training" (default) or "testing"
    if dataset == "training":
        filename_digits = os.path.join(path, 'train-images-idx3-ubyte')
        filename_labels = os.path.join(path, 'train-labels-idx1-ubyte')
    elif dataset == "testing":
        filename_digits = os.path.join(path, 't10k-images-idx3-ubyte')
        filename_labels = os.path.join(path, 't10k-labels-idx1-ubyte')
    else:
        raise ValueError("dataset must be 'testing' or 'training'")

    # Import digits data
    digits_file_object = open(filename_digits, 'rb')
    magic_nr, size, rows, cols = struct.unpack(">IIII", digits_file_object.read(16))
    digits_data = array("B", digits_file_object.read())
    digits_file_object.close()

    # Import label data
    labels_file_object = open(filename_labels, 'rb')
    magic_nr, size = struct.unpack(">II", labels_file_object.read(8))
    labels_data = array("B", labels_file_object.read())
    labels_file_object.close()

    # Find indices of selected digits
    indices = [k for k in range(size) if labels_data[k] in selected_digits]
    n = len(indices)

    # Create empty arrays for vector X and T
    X = np.zeros((n, rows * cols), dtype=np.uint8)
    T = np.zeros((n, 1), dtype=np.uint8)

    # Fill vector X from digits_data
    # Fill vector T from labels_data
    for i in range(n):
        X[i] = digits_data[indices[i] * rows * cols:(indices[i] + 1) * rows * cols]
        T[i] = labels_data[indices[i]]

    return X, T


def vector_to_img(*args, **kwargs):
    print kwargs
    # plt.imshow(v.reshape(28, 28), interpolation='None', cmap='gray')
    # plt.axis('off')
    # if show:
    #    plt.show()
    width = kwargs["width"] if kwargs.has_key("width") else 28
    size = 1
    n = len(args)
    fig = plt.figure()
    for i, arg in enumerate(args):
        plt.subplot(1,n,i+1)
        plt.imshow(arg.reshape(width, width), interpolation="None", cmap="gray")
        plt.axis("off")
    fig.tight_layout(pad=0)
    fig.set_size_inches(w=n*size,h=size)
    plt.show()


def PCA(X, U, V):
    # Get recentered matrix Z
    Z = X - U

    # Get principal component matrix
    P = np.dot(Z, V.T)
    # print P
    # scatter_plot(P, T)

    return P


def XZCVPR(X, T, selected_digits):
    # vector_to_img(X[4], X[-1])

    # Get mean matrix U and matrix Z
    U = np.mean(X, axis=0)
    # print U
    # print type(U)
    # print U.shape

    # Get recentered matrix Z
    Z = X - U
    # print Z
    # print np.mean(Z, axis=0)

    # Get covariance matrix C
    C = np.cov(Z, rowvar=False)
    # print C
    # vector_to_img(C, width=768)

    L, V = LA.eigh(C)
    # print L
    # print V

    # row = V[-1, :]
    # col = V[:, -1]
    # print np.dot(C, row) - (L[-1] * row)
    # print np.dot(C, col) - (L[-1] * col)

    L = np.flipud(L)
    V = np.flipud(V.T)
    # row = V[0, :]
    # print np.dot(C, row) - (L[0] * row)
    # print np.allclose(np.dot(C, row), L[0] * row)
    # vector_to_img(V[0:1, :], width=20)
    # vector_to_img(V[1:2, :], width=20)

    # Get principal component matrix
    P = np.dot(Z, V.T)
    # print P
    # scatter_plot(P[:, 0:2], T, selected_digits=selected_digits)

    # Get recovered matrix
    R = np.dot(P, V)
    # print R-Z

    # Get recovered data matrix
    Xrec = R+U
    # print Xrec-X

    # reduced X
    Xrec = (np.dot(P[:, 0:2], V[0:2, :])) + U
    # print Xrec
    # vector_to_img(X[0], Xrec[0])

    return U, V, P, L


def scatter_plot(P, T, selected_digits=[5, 6], pngname=None):
    # For best effect, points should not be drawn in sequence but in random order
    np.random.seed(0)
    randomorder = np.random.permutation(np.arange(len(T)))
    randomorder = np.arange(len(T))

    # Set colors
    cols = np.zeros((len(T), 4))  # Initialize matrix to hold colors
    for i in range(len(T)):
        if T[i] == selected_digits[0]:
            cols[i] = [1, 0, 0, 0.25]  # Negative points are red (with opacity 0.25)
        else:
            cols[i] = [0, 1, 0, 0.25]  # Positive points are green (with opacity 0.25)

    # Draw scatter plot
    fig = plt.figure()
    ax = fig.add_subplot(111, facecolor='black')
    ax.scatter(P[randomorder, 1], P[randomorder, 0], s=5, linewidths=0, facecolors=cols[randomorder, :], marker="o")
    ax.set_aspect('equal')

    plt.gca().invert_yaxis()
    if pngname:
        plt.savefig(pngname)
    else:
        plt.show()

def build_2d_histograms(Prec, T, B, selected_digits):
    histograms = [np.zeros((B, B), dtype=np.int), np.zeros((B, B), dtype=np.int32)]

    mins = np.amin(Prec, axis=0)
    # print "mins: ", mins

    maxs = np.amax(Prec, axis=0)
    # print "maxs: ", maxs

    for i in range(len(Prec)):
        r = int(np.round(np.divide((B - 1) * (Prec[i, 0] - mins[0]), maxs[0] - mins[0], dtype=np.float64)))
        c = int(np.round(np.divide((B - 1) * (Prec[i, 1] - mins[1]), maxs[1] - mins[1], dtype=np.float64)))
        if T[i] == selected_digits[0]:
            histograms[0][r, c] += 1
        else:
            histograms[1][r, c] += 1
    # print histograms
    # vector_to_img(histograms[0], width=B)
    # vector_to_img(histograms[1], width=B)

    return histograms, maxs, mins


def get_ccps_by_2d_histograms(B, histograms, maxs, mins, n, x):
    r = int(np.round(np.divide((B - 1) * (x[0] - mins[0]), maxs[0] - mins[0], dtype=np.float64)))
    c = int(np.round(np.divide((B - 1) * (x[1] - mins[1]), maxs[1] - mins[1], dtype=np.float64)))

    ccp = [0., 0.]
    ccp[0] = np.divide(histograms[0][r, c], n[0], dtype=np.float64)
    ccp[1] = np.divide(histograms[1][r, c], n[1], dtype=np.float64)

    return ccp


def build_2d_beysian(Prec, T, selected_digits):
    mus = [None, None]
    mus[0] = np.mean(Prec[T[:, 0] == selected_digits[0]], axis=0)
    mus[1] = np.mean(Prec[T[:, 0] == selected_digits[1]], axis=0)

    Sigmas = [None, None]
    Sigmas[0] = np.cov(Prec[T[:, 0] == selected_digits[0]], rowvar=False)
    Sigmas[1] = np.cov(Prec[T[:, 0] == selected_digits[1]], rowvar=False)

    return mus, Sigmas


def get_ccp_by_2d_beysian(x, mu, Sigma):
    # print mu
    d = np.alen(mu)
    # print d
    dfact1 = (2 * np.pi) ** d
    dfact2 = np.linalg.det(Sigma)
    fact = np.divide(1., np.sqrt(dfact1 * dfact2))
    xc = x - mu
    Invsigma = np.linalg.inv(Sigma)
    power = -0.5 * np.dot(np.dot(xc, Invsigma), xc.T)
    return fact * np.exp(power)


def get_ccps_by_2d_beysian(x, mus, Sigmas):
    return [get_ccp_by_2d_beysian(x, mus[0], Sigmas[0]), get_ccp_by_2d_beysian(x, mus[1], Sigmas[1])]


def posterior_probability(ccps, n):
    total = sum(n)
    # print n[0], n[1], total

    eachps = [0.] * len(n)
    # print eachps
    for i in range(len(n)):
        eachps[i] = ccps[i] * np.divide(n[i], total, dtype=np.float64)
    # print eachps

    ttlp = sum(eachps)
    # print ttlp

    pps = [0.] * len(n)
    # print pps
    for i in range(len(n)):
        if ttlp == 0.:
            pps[i] = 0.
        else:
            pps[i] = np.divide(eachps[i], ttlp, dtype=np.float64)
    # print pps

    return pps


def main():
    selected_digits = ["male", "female"]

    excelfile = "/Users/ChiYuChen/Intro to Machine Learning and Data Mining/Final Project/Assignment_3_Submission_Template_Original.xlsx"
    sheetname = "Training Data"
    X = np.array(readExcel(excelfile, sheetname=sheetname, startrow=2, endrow=3169, endcol=20), dtype=np.float64)
    T = np.array(readExcel(excelfile, sheetname=sheetname, startrow=2, endrow=3169, startcol=21, endcol=21),
                  dtype=np.string_)

    # X, T = load_mnist("training", selected_digits)
    # X = np.array(X, np.float64)
    # T = np.array(T, np.float64)

    # print type(X)
    # print type(T)
    # print X.shape
    # print T.shape

    N = X.shape[0]
    # print N

    # Xn = np.array([[72., 101., 94.], [50., 96., 70.], [14., 79., 10.], [8., 70., 1.]], np.float64)
    U, V, P, L = XZCVPR(X, T, selected_digits)
    print L
    print V

    n = [0, 0]
    for i in range(len(T)):
        if T[i][0] == selected_digits[0]:
            n[0] += 1
        else:
            n[1] += 1
    # n[0] = len(T[T == selected_digits[0]])
    # n[1] = len(T[T == selected_digits[1]])
    # print n

    B = int(np.ceil(np.log2(n[0] + n[1]))) + 1
    # B = 32
    # print B

    # print P
    print len(P[0])
    mxcb = (0, 1)
    mxacc = 0.
    mxhistograms = None
    mxmaxs = None
    mxmins = None
    mxmus = None
    mxSigmas = None
    mxbccmh = None
    mxbccmb = None
    mxmetricsh = None
    mxmetricsb = None
    acctable = dict()
    acclistk = []
    acclisth = []
    acclistb = []
    for sidx in range(len(P[0])-1):
        for eidx in range(sidx+1, len(P[0])):
            # print sidx, eidx
            Prec = P[:, [sidx, eidx]]
            # print Prec

            histograms, maxs, mins = build_2d_histograms(Prec, T, B, selected_digits)

            mus, Sigmas = build_2d_beysian(Prec, T, selected_digits)

            # Th = np.zeros((len(T), 2))
            # Tb = np.zeros((len(T), 2))
            Th = [None] * len(T)
            Tb = [None] * len(T)
            for i in range(len(T)):
                ccps_by_histograms = get_ccps_by_2d_histograms(B, histograms, maxs, mins, n, Prec[i])
                # print ccps_by_histograms
                pps_by_histograms = posterior_probability(ccps_by_histograms, n)
                # print pps_by_histograms
                if pps_by_histograms[0] == pps_by_histograms[1]:
                    Th[i] = ("Undecidable", 0.)
                else:
                    if pps_by_histograms[0] > pps_by_histograms[1]:
                        Th[i] = (selected_digits[0], pps_by_histograms[0])
                    else:
                        Th[i] = (selected_digits[1], pps_by_histograms[1])

                ccps_by_beysian = get_ccps_by_2d_beysian(Prec[i], mus, Sigmas)
                # print ccps_by_beysian
                pps_by_beysian = posterior_probability(ccps_by_beysian, n)
                # print pps_by_beysian
                if pps_by_beysian[0] == pps_by_beysian[1]:
                    Tb[i] = ("Undecidable", 0.)
                else:
                    if pps_by_beysian[0] > pps_by_beysian[1]:
                        Tb[i] = (selected_digits[0], pps_by_beysian[0])
                    else:
                        Tb[i] = (selected_digits[1], pps_by_beysian[1])

            # print Th
            # print Tb
            # print len(T[T == Th])
            # print len(T[T == Tb])

            bccmh = np.zeros((2, 2), dtype=np.int32)
            for i in range(len(T)):
                row = 1 if T[i][0] == selected_digits[1] else 0
                col = 1 if Th[i][0] == selected_digits[1] else 0
                # print row, col
                bccmh[row, col] += 1
            # print bccmh

            metricsh = [
                np.divide(bccmh[0, 0] + bccmh[1, 1], bccmh[0, 0] + bccmh[0, 1] + bccmh[1, 0] + bccmh[1, 1], dtype=np.float64),
                np.divide(bccmh[1, 1], bccmh[1, 0] + bccmh[1, 1], dtype=np.float64),
                np.divide(bccmh[0, 0], bccmh[0, 0] + bccmh[0, 1], dtype=np.float64),
                np.divide(bccmh[1, 1], bccmh[0, 1] + bccmh[1, 1], dtype=np.float64)
            ]
            # print metricsh

            bccmb = np.zeros((2, 2), dtype=np.int32)
            for i in range(len(T)):
                row = 1 if T[i][0] == selected_digits[1] else 0
                col = 1 if Tb[i][0] == selected_digits[1] else 0
                # print row, col
                bccmb[row, col] += 1
            # print bccmb

            metricsb = [
                np.divide(bccmb[0, 0] + bccmb[1, 1], bccmb[0, 0] + bccmb[0, 1] + bccmb[1, 0] + bccmb[1, 1],
                          dtype=np.float64),
                np.divide(bccmb[1, 1], bccmb[1, 0] + bccmb[1, 1], dtype=np.float64),
                np.divide(bccmb[0, 0], bccmb[0, 0] + bccmb[0, 1], dtype=np.float64),
                np.divide(bccmb[1, 1], bccmb[0, 1] + bccmb[1, 1], dtype=np.float64)
            ]
            # print metricsb

            ch = 0
            cb = 0
            for i in range(len(T)):
                # print Th[i][0], T[i][0], Th[i][0] == T[i][0]
                if Th[i][0] == T[i][0]:
                    ch += 1
                # print Tb[i][0], T[i][0], Tb[i][0] == T[i][0]
                if Tb[i][0] == T[i][0]:
                    cb += 1
            # print len(T), ch, cb

            accuracyh = np.divide(ch, len(T), dtype=np.float64)
            accuracyb = np.divide(cb, len(T), dtype=np.float64)
            # print accuracyh
            # print accuracyb

            key = "v%d, v%d" % (sidx, eidx)
            acctable[key] = (accuracyh, accuracyb)
            acclistk.append(key)
            acclisth.append(accuracyh)
            acclistb.append(accuracyb)

            if (accuracyh + accuracyb) > mxacc:
                mxacc = (accuracyh + accuracyb)
                mxcb = (sidx, eidx)
                mxhistograms = histograms
                mxmaxs = maxs
                mxmins = mins
                mxmus = mus
                mxSigmas = Sigmas
                mxbccmh = bccmh
                mxbccmb = bccmb
                mxmetricsh = metricsh
                mxmetricsb = metricsb

                print sidx, eidx, accuracyh, accuracyb
                print Prec
                Prec_nl = np.zeros((len(T), 2), dtype=np.float64)
                for r in range(len(T)):
                    Prec_nl[r, 0] = np.divide((Prec[r, 0] - mins[0]), maxs[0] - mins[0], dtype=np.float64)
                    Prec_nl[r, 1] = np.divide((Prec[r, 1] - mins[1]), maxs[1] - mins[1], dtype=np.float64)
                # print histograms
                scatter_plot(Prec_nl, T, selected_digits=selected_digits, pngname="p%d-p%d.png" % (sidx, eidx))

    print "Start to output the result to excel file"

    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook

    excelfile = "/Users/ChiYuChen/Intro to Machine Learning and Data Mining/Final Project/Assignment_3_Submission_Template.xlsx"
    sheetname = "Results"

    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # print [U]
    df = DataFrame([U])
    df.to_excel(writer, sheet_name=sheetname, startrow=1, startcol=1, header=False, index=False)

    for i in range(len(P[0])):
        # print [[L[i]]]
        df = DataFrame([[L[i]]])
        df.to_excel(writer, sheet_name=sheetname, startrow=(i + 3), startcol=1, header=False, index=False)

        # print [V[i, :]]
        df = DataFrame([V[i, :]])
        df.to_excel(writer, sheet_name=sheetname, startrow=(i + 3), startcol=3, header=False, index=False)

    # print [n[0]]
    df = DataFrame([n[0]])
    df.to_excel(writer, sheet_name=sheetname, startrow=24, startcol=1, header=False, index=False)

    # print [n[1]]
    df = DataFrame([n[1]])
    df.to_excel(writer, sheet_name=sheetname, startrow=25, startcol=1, header=False, index=False)

    # print [acclistk]
    df = DataFrame([acclistk])
    # print [acctable.keys()]
    # df = DataFrame([acctable.keys()])
    df.to_excel(writer, sheet_name=sheetname, startrow=27, startcol=1, header=False, index=False)
    # print [acclisth]
    df = DataFrame([acclisth])
    # print accuracyh
    # df = DataFrame([[acc[0] for acc in acctable.values()]])
    df.to_excel(writer, sheet_name=sheetname, startrow=28, startcol=1, header=False, index=False)
    # print [acclistb]
    df = DataFrame([acclistb])
    # print accuracyb
    # df = DataFrame([[acc[1] for acc in acctable.values()]])
    df.to_excel(writer, sheet_name=sheetname, startrow=29, startcol=1, header=False, index=False)

    # print [[mxcb[0], mxcb[1]]]
    df = DataFrame([[mxcb[0], mxcb[1]]])
    df.to_excel(writer, sheet_name=sheetname, startrow=31, startcol=1, header=False, index=False)

    # print mxbccmh
    df = DataFrame(mxbccmh)
    df.to_excel(writer, sheet_name=sheetname, startrow=35, startcol=3, header=False, index=False)

    # print mxmetricsh
    df = DataFrame(mxmetricsh)
    df.to_excel(writer, sheet_name=sheetname, startrow=33, startcol=7, header=False, index=False)

    # print mxbccmb
    df = DataFrame(mxbccmb)
    df.to_excel(writer, sheet_name=sheetname, startrow=40, startcol=3, header=False, index=False)

    # print mxmetricsb
    df = DataFrame(mxmetricsb)
    df.to_excel(writer, sheet_name=sheetname, startrow=38, startcol=7, header=False, index=False)

    # print [mxmus[0]]
    df = DataFrame([mxmus[0]])
    df.to_excel(writer, sheet_name=sheetname, startrow=43, startcol=1, header=False, index=False)

    # print [mxmus[1]]
    df = DataFrame([mxmus[1]])
    df.to_excel(writer, sheet_name=sheetname, startrow=44, startcol=1, header=False, index=False)

    # print mxSigmas[0]
    df = DataFrame(mxSigmas[0])
    df.to_excel(writer, sheet_name=sheetname, startrow=46, startcol=1, header=False, index=False)

    # print mxSigmas[1]
    df = DataFrame(mxSigmas[1])
    df.to_excel(writer, sheet_name=sheetname, startrow=48, startcol=1, header=False, index=False)

    # print mxmaxs[0], mxmins[0]
    df = DataFrame([[mxmaxs[0], mxmins[0]]])
    df.to_excel(writer, sheet_name=sheetname, startrow=51, startcol=1, header=False, index=False)

    # print mxmaxs[1], mxmins[1]
    df = DataFrame([[mxmaxs[1], mxmins[1]]])
    df.to_excel(writer, sheet_name=sheetname, startrow=52, startcol=1, header=False, index=False)

    # print mxhistograms[0]
    # print mxhistograms[0][0, :]
    df = DataFrame(mxhistograms[0])
    df.to_excel(writer, sheet_name=sheetname, startrow=54, startcol=1, header=False, index=False)

    # print mxhistograms[1]
    # print mxhistograms[1][0, :]
    df = DataFrame(mxhistograms[1])
    df.to_excel(writer, sheet_name=sheetname, startrow=68, startcol=1, header=False, index=False)

    writer.save()
    writer.close()


if __name__ == "__main__":
    main()
