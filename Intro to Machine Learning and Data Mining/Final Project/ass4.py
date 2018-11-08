import os
import struct
import numpy as np
from numpy import linalg as LA
from array import array
import matplotlib.pyplot as plt


def readExcelSheet1(excelfile):
    from pandas import read_excel
    return (read_excel(excelfile)).values


def readExcelRange(excelfile, sheetname="Sheet1", startrow=1, endrow=1, startcol=1, endcol=1):
    from pandas import read_excel
    values = (read_excel(excelfile, sheetname, header=None)).values
    return values[startrow-1:endrow, startcol-1:endcol]


def readExcel(excelfile, **args):
    if args:
        data = readExcelRange(excelfile, **args)
    else:
        data = readExcelSheet1(excelfile)
    if data.shape == (1, 1):
        return data[0, 0]
    elif (data.shape)[0] == 1:
        return data[0]
    else:
        return data


def getSheetNames(excelfile):
    from pandas import ExcelFile
    return (ExcelFile(excelfile)).sheet_names


def writeExcelData(x, excelfile, sheetname, startrow, startcol, columns):
    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook
    df = DataFrame(x)
    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheetname, columns=columns, startrow=startrow-1, startcol=startcol-1, header=False, index=False)
    writer.save()
    writer.close()


def load_mnist(dataset, selected_digits, path=os.path.dirname(os.path.realpath(__file__))):

    # Check training/testing specification. Must be "training" (default) or "testing"
    if dataset == "training":
        filename_digits = os.path.join(path, 'train-images-idx3-ubyte')
        filename_labels = os.path.join(path, 'train-labels-idx1-ubyte')
    elif dataset == "testing":
        filename_digits = os.path.join(path, 't10k-images-idx3-ubyte')
        filename_labels = os.path.join(path, 't10k-labels-idx1-ubyte')
    else:
        raise ValueError("dataset must be 'testing' or 'training'")

    # Import digits data
    digits_file_object = open(filename_digits, 'rb')
    magic_nr, size, rows, cols = struct.unpack(">IIII", digits_file_object.read(16))
    digits_data = array("B", digits_file_object.read())
    digits_file_object.close()

    # Import label data
    labels_file_object = open(filename_labels, 'rb')
    magic_nr, size = struct.unpack(">II", labels_file_object.read(8))
    labels_data = array("B", labels_file_object.read())
    labels_file_object.close()

    # Find indices of selected digits
    indices = [k for k in range(size) if labels_data[k] in selected_digits]
    n = len(indices)

    # Create empty arrays for vector X and T
    X = np.zeros((n, rows * cols), dtype=np.uint8)
    T = np.zeros((n, 1), dtype=np.uint8)

    # Fill vector X from digits_data
    # Fill vector T from labels_data
    for i in range(n):
        X[i] = digits_data[indices[i] * rows * cols:(indices[i] + 1) * rows * cols]
        T[i] = labels_data[indices[i]]

    return X, T


def vector_to_img(*args, **kwargs):
    print kwargs
    # plt.imshow(v.reshape(28, 28), interpolation='None', cmap='gray')
    # plt.axis('off')
    # if show:
    #    plt.show()
    width = kwargs["width"] if kwargs.has_key("width") else 28
    size = 1
    n = len(args)
    fig = plt.figure()
    for i, arg in enumerate(args):
        plt.subplot(1,n,i+1)
        plt.imshow(arg.reshape(width, width), interpolation="None", cmap="gray")
        plt.axis("off")
    fig.tight_layout(pad=0)
    fig.set_size_inches(w=n*size,h=size)
    plt.show()


def PCA(X, U, V):
    # Get recentered matrix Z
    Z = X - U

    # Get principal component matrix
    P = np.dot(Z, V.T)
    # print P
    # scatter_plot(P, T)

    return P


def XZCVPR(X, T, selected_digits):
    # vector_to_img(X[4], X[-1])

    # Get mean matrix U and matrix Z
    U = np.mean(X, axis=0)
    # print U
    # print type(U)
    # print U.shape

    # Get recentered matrix Z
    Z = X - U
    # print Z
    # print np.mean(Z, axis=0)

    # Get covariance matrix C
    C = np.cov(Z, rowvar=False)
    # print C
    # vector_to_img(C, width=784)

    L, V = LA.eigh(C)
    # print L
    # print V

    # row = V[-1, :]
    # col = V[:, -1]
    # print np.dot(C, row) - (L[-1] * row)
    # print np.dot(C, col) - (L[-1] * col)

    L = np.flipud(L)
    V = np.flipud(V.T)
    # row = V[0, :]
    # print np.dot(C, row) - (L[0] * row)
    # print np.allclose(np.dot(C, row), L[0] * row)
    # vector_to_img(V[0:1, :], width=28)
    # vector_to_img(V[1:2, :], width=28)

    # Get principal component matrix
    P = np.dot(Z, V.T)
    # print P
    scatter_plot(P, T, selected_digits=selected_digits)

    # Get recovered matrix
    R = np.dot(P, V)
    # print R-Z

    # Get recovered data matrix
    Xrec = R+U
    # print Xrec-X

    # reduced X
    Xrec = (np.dot(P[:, 0:2], V[0:2, :])) + U
    # print Xrec
    # vector_to_img(X[0], Xrec[0])

    return U, V, P


def scatter_plot(P, T, selected_digits=[5, 6]):
    # For best effect, points should not be drawn in sequence but in random order
    np.random.seed(0)
    randomorder = np.random.permutation(np.arange(len(T)))
    randomorder = np.arange(len(T))

    # Set colors
    cols = np.zeros((len(T), 4))  # Initialize matrix to hold colors
    for i in range(len(T)):
        if T[i] == selected_digits[0]:
            cols[i] = [1, 0, 0, 0.25]  # Negative points are red (with opacity 0.25)
        else:
            cols[i] = [0, 1, 0, 0.25]  # Positive points are green (with opacity 0.25)

    # Draw scatter plot
    fig = plt.figure()
    ax = fig.add_subplot(111, facecolor='black')
    ax.scatter(P[randomorder, 1], P[randomorder, 0], s=5, linewidths=0, facecolors=cols[randomorder, :], marker="o");
    ax.set_aspect('equal')

    plt.gca().invert_yaxis()
    plt.show()


def build_2d_histograms(Prec, T, B, selected_digits):
    histograms = [np.zeros((B, B), dtype=np.int), np.zeros((B, B), dtype=np.int32)]

    mins = np.amin(Prec, axis=0)
    # print "mins: ", mins

    maxs = np.amax(Prec, axis=0)
    # print "maxs: ", maxs

    for i in range(len(Prec)):
        r = int(np.round(np.divide((B - 1) * (Prec[i, 0] - mins[0]), maxs[0] - mins[0], dtype=np.float64)))
        c = int(np.round(np.divide((B - 1) * (Prec[i, 1] - mins[1]), maxs[1] - mins[1], dtype=np.float64)))
        if T[i] == selected_digits[0]:
            histograms[0][r, c] += 1
        else:
            histograms[1][r, c] += 1
    # print histograms
    vector_to_img(histograms[0], width=B)
    vector_to_img(histograms[1], width=B)

    return histograms, maxs, mins


def get_ccps_by_2d_histograms(B, histograms, maxs, mins, n, x):
    r = int(np.round(np.divide((B - 1) * (x[0] - mins[0]), maxs[0] - mins[0], dtype=np.float64)))
    c = int(np.round(np.divide((B - 1) * (x[1] - mins[1]), maxs[1] - mins[1], dtype=np.float64)))

    ccp = [0., 0.]
    ccp[0] = np.divide(histograms[0][r, c], n[0], dtype=np.float64)
    ccp[1] = np.divide(histograms[1][r, c], n[1], dtype=np.float64)

    return ccp


def build_2d_beysian(Prec, T, selected_digits):
    mus = [None, None]
    mus[0] = np.mean(Prec[T[:, 0] == selected_digits[0]], axis=0)
    mus[1] = np.mean(Prec[T[:, 0] == selected_digits[1]], axis=0)

    Sigmas = [None, None]
    Sigmas[0] = np.cov(Prec[T[:, 0] == selected_digits[0]], rowvar=False)
    Sigmas[1] = np.cov(Prec[T[:, 0] == selected_digits[1]], rowvar=False)

    return mus, Sigmas


def get_ccp_by_2d_beysian(x, mu, Sigma):
    # print mu
    d = np.alen(mu)
    # print d
    dfact1 = (2 * np.pi) ** d
    dfact2 = np.linalg.det(Sigma)
    fact = np.divide(1., np.sqrt(dfact1 * dfact2))
    xc = x - mu
    Invsigma = np.linalg.inv(Sigma)
    power = -0.5 * np.dot(np.dot(xc, Invsigma), xc.T)
    return fact * np.exp(power)


def get_ccps_by_2d_beysian(x, mus, Sigmas):
    return [get_ccp_by_2d_beysian(x, mus[0], Sigmas[0]), get_ccp_by_2d_beysian(x, mus[1], Sigmas[1])]


def posterior_probability(ccps, n):
    total = sum(n)
    # print n[0], n[1], total

    eachps = [0.] * len(n)
    # print eachps
    for i in range(len(n)):
        eachps[i] = ccps[i] * np.divide(n[i], total, dtype=np.float64)
    # print eachps

    ttlp = sum(eachps)
    # print ttlp

    pps = [0.] * len(n)
    # print pps
    for i in range(len(n)):
        if ttlp == 0.:
            pps[i] = 0.
        else:
            pps[i] = np.divide(eachps[i], ttlp, dtype=np.float64)
    # print pps

    return pps


def main():
    excelfile = "/Users/ChiYuChen/Intro to Machine Learning and Data Mining/Final Project/Final_Project_Data_and_Template_Original.xlsx"

    sheetname = "Training Data"
    # df = readExcel(excelfile, sheetname=sheetname, startrow=2, endrow=6601, endcol=17)
    # print df
    # print df[0]
    # print df[-1]

    X = np.array(readExcel(excelfile, sheetname=sheetname, startrow=2, endrow=3169, endcol=20), dtype=np.float64)
    print X

    Xa = np.insert(X, 0, 1., axis=1)
    print Xa

    Tw = np.array(readExcel(excelfile, sheetname=sheetname, startrow=2, endrow=3169, startcol=21, endcol=21), dtype=np.string_)
    print Tw

    T = np.array([1 if Tw[i][0] == "male" else -1 for i in range(len(Tw))], dtype=np.int32)
    print T

    Xapi = np.linalg.pinv(Xa)
    print Xapi

    W = np.dot(Xapi, T)
    print W

    T_validating = np.dot(Xa, W)
    print T_validating
    T_validating = np.array([1 if T_validating[i] > 0 else -1 for i in range(len(T_validating))], dtype=np.int32)
    print T_validating

    bccm = np.zeros((2, 2), dtype=np.int32)
    for i in range(len(T)):
        row = 1 if T[i] == 1 else 0
        col = 1 if T_validating[i] == 1 else 0
        # print row, col
        bccm[row, col] += 1
    print bccm

    metrics = [
        np.divide(bccm[0, 0] + bccm[1, 1], bccm[0, 0] + bccm[0, 1] + bccm[1, 0] + bccm[1, 1], dtype=np.float64),
        np.divide(bccm[1, 1], bccm[1, 0] + bccm[1, 1], dtype=np.float64),
        np.divide(bccm[0, 0], bccm[0, 0] + bccm[0, 1], dtype=np.float64),
        np.divide(bccm[1, 1], bccm[0, 1] + bccm[1, 1], dtype=np.float64)
    ]
    print metrics

    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook

    excelfile = "/Users/ChiYuChen/Intro to Machine Learning and Data Mining/Final Project/Final_Project_Data_and_Template.xlsx"
    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    sheetname = "Classifiers"

    # print W
    df = DataFrame(W)
    df.to_excel(writer, sheet_name=sheetname, startrow=4, startcol=0, header=False, index=False)

    sheetname = "Performance"

    # print bccm
    df = DataFrame(bccm)
    df.to_excel(writer, sheet_name=sheetname, startrow=9, startcol=2, header=False, index=False)

    # print metrics
    df = DataFrame(metrics)
    df.to_excel(writer, sheet_name=sheetname, startrow=7, startcol=6, header=False, index=False)

    writer.save()
    writer.close()


if __name__ == "__main__":
    main()
